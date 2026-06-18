"""
merge_new_invoices.py
─────────────────────
Drop new invoice files into:  Raw Data/New Invoices/
Then run (from the repo root):  python automation/merge_new_invoices.py

What it does:
  1. Auto-detects airline from each file's content
  2. Parses the file into standardised rows
  3. Deduplicates against existing Sullivan_Financials.xlsx
  4. Appends new rows, saves the Excel
  5. Moves processed files into Raw Data/New Invoices/Processed/
  6. Prints a plain-English summary

Supported formats:
  VS  — xlsx invoice detail report (per-line-item, 30K+ rows)
  ET  — xlsx with AR Invoice sheet (per-flight food/service split)
  SK  — CSV invoice export (per-line-item, aggregated by invoice #)
  AA  — xlsx Sage Intacct export (per-invoice, food/service split)
"""

import sys, shutil, re
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Paths ─────────────────────────────────────────────────────────────────────
ROOT          = Path(__file__).parent.parent
DROP_DIR      = ROOT / "Raw Data" / "New Invoices"
PROCESSED_DIR = DROP_DIR / "Processed"
EXCEL_FILE    = ROOT / "Sullivan_Financials.xlsx"

PROCESSED_DIR.mkdir(parents=True, exist_ok=True)

# ── Column layout in the Flights sheet ───────────────────────────────────────
# A  B     C                  D             E               F             G         H       I      J
# Al Date  Flight/Invoice     Food Revenue  Service Revenue Total Revenue Pax Count Status  Notes  Source File
# K (Unique Key formula)  L (Duplicate? formula)
FLIGHTS_SHEET = "Flights"
HEADER_ROW    = 2   # row 1 = instruction banner, row 2 = column headers


def money_fmt():  return '$#,##0.00;($#,##0.00);"-"'
def thin_border():
    s = Side(style='thin', color='CBD5E1')
    return Border(left=s, right=s, top=s, bottom=s)
def fill(hex_color):
    return PatternFill('solid', start_color=hex_color, fgColor=hex_color)

AIRLINE_COLORS = {'VS': 'FFF5F5', 'AA': 'F0F5FF', 'ET': 'F0FFF4', 'SK': 'F5F0FF'}


# ══════════════════════════════════════════════════════════════════════════════
# AIRLINE DETECTION
# ══════════════════════════════════════════════════════════════════════════════
def detect_airline(path: Path) -> str | None:
    ext  = path.suffix.lower()
    name = path.name.lower()

    if ext == '.csv':
        try:
            sample = path.read_text(errors='ignore')[:4000].lower()
        except Exception:
            return None
        if 'sk-' in sample or 'brd flt' in sample or 'srv. flt' in sample:
            return 'SK'
        return None

    if ext == '.xlsx':
        try:
            xl     = pd.ExcelFile(path, engine='openpyxl')
            sheets = [s.lower() for s in xl.sheet_names]

            # ET always has an AR Invoice sheet
            if 'ar invoice' in sheets:
                return 'ET'

            # VS: huge file with per-line-item rows, look for V104/V110/V116
            df0 = xl.parse(xl.sheet_names[0], nrows=5, dtype=str)
            col_str = ' '.join(str(c).lower() for c in df0.columns)
            if ('flight no' in col_str or 'flight date' in col_str
                    or 'meal type' in col_str or 'invoiced pax' in col_str):
                return 'VS'

            # AA: Sage per-invoice export (food + service columns)
            if 'food' in col_str or 'service' in col_str:
                return 'AA'

        except Exception as e:
            print(f"  ⚠  Could not inspect {path.name}: {e}")
        return None

    return None


# ══════════════════════════════════════════════════════════════════════════════
# PARSERS  — each returns a list of dicts with standard keys
# ══════════════════════════════════════════════════════════════════════════════
STD_KEYS = ['Airline', 'Date', 'Flight / Invoice',
            'Food Revenue', 'Service Revenue', 'Total Revenue',
            'Pax Count', 'Status', 'Notes', 'Source File']


def _to_date(v) -> str:
    """Return YYYY-MM-DD string or '' if unparseable."""
    if pd.isna(v) or v is None:
        return ''
    try:
        d = pd.to_datetime(v)
        return d.strftime('%Y-%m-%d')
    except Exception:
        return str(v)


def _money(v) -> float | None:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    try:
        return float(str(v).replace('$', '').replace(',', '').strip())
    except Exception:
        return None


# ─── Virgin Atlantic ──────────────────────────────────────────────────────────
def parse_vs(path: Path) -> list[dict]:
    print("  Parsing as Virgin Atlantic (VS) invoice detail…")
    df = pd.read_excel(path, dtype=str)

    # Find column names flexibly
    def find_col(keywords):
        for c in df.columns:
            if all(k.lower() in c.lower() for k in keywords):
                return c
        return None

    date_col   = find_col(['flight', 'date']) or find_col(['date'])
    flight_col = find_col(['flight', 'no'])   or find_col(['flt'])
    cost_col   = find_col(['tot', 'cost'])    or find_col(['total'])
    pax_col    = find_col(['actual', 'pax'])  or find_col(['invoiced', 'pax'])
    status_col = find_col(['invoice', 'status'])

    if not date_col or not cost_col:
        print("  ✗  Could not find required columns in VS file — skipping.")
        return []

    df[cost_col]  = pd.to_numeric(df[cost_col].str.replace(r'[$,]', '', regex=True), errors='coerce')
    df[date_col]  = pd.to_datetime(df[date_col], errors='coerce')
    if pax_col:
        df[pax_col] = pd.to_numeric(df[pax_col], errors='coerce')

    group_cols = [date_col]
    if flight_col:
        group_cols.append(flight_col)

    agg = {cost_col: 'sum'}
    if pax_col:
        agg[pax_col] = 'max'
    grouped = df.groupby(group_cols).agg(agg).reset_index()

    rows = []
    for _, r in grouped.iterrows():
        date  = _to_date(r[date_col])
        fi    = str(r[flight_col]).strip() if flight_col else 'VS-FLIGHT'
        total = _money(r[cost_col])
        pax   = int(r[pax_col]) if pax_col and not pd.isna(r[pax_col]) else None
        rows.append({
            'Airline': 'VS', 'Date': date, 'Flight / Invoice': fi,
            'Food Revenue': None, 'Service Revenue': None, 'Total Revenue': total,
            'Pax Count': pax, 'Status': 'Invoiced',
            'Notes': 'Aggregated from VS raw invoice detail (food/svc split via Sage Intacct)',
            'Source File': path.name,
        })
    return rows


# ─── Ethiopian Airlines ───────────────────────────────────────────────────────
def parse_et(path: Path) -> list[dict]:
    print("  Parsing as Ethiopian Airlines (ET) AR Invoice…")
    try:
        df = pd.read_excel(path, sheet_name='AR Invoice', dtype=str)
    except Exception as e:
        print(f"  ✗  Could not read AR Invoice sheet: {e}")
        return []

    def find_col(keywords):
        for c in df.columns:
            if all(k.lower() in c.lower() for k in keywords):
                return c
        return None

    date_col    = find_col(['date']) or find_col(['flight', 'date'])
    flight_col  = find_col(['flight']) or find_col(['flt'])
    food_col    = find_col(['food'])   or find_col(['411001'])
    svc_col     = find_col(['service']) or find_col(['413909'])
    total_col   = find_col(['total'])
    biz_col     = find_col(['business', 'pax']) or find_col(['biz'])
    econ_col    = find_col(['economy', 'pax']) or find_col(['econ'])

    rows = []
    for _, r in df.iterrows():
        date = _to_date(r.get(date_col, ''))
        if not date or date < '2020-01-01':
            continue   # skip template/header rows with bogus dates

        fi    = str(r.get(flight_col, 'ET519')).strip() or 'ET519'
        food  = _money(r.get(food_col))
        svc   = _money(r.get(svc_col))
        total = _money(r.get(total_col))
        if total is None and food is not None and svc is not None:
            total = food + svc

        biz   = _money(r.get(biz_col))
        econ  = _money(r.get(econ_col))
        pax   = int(biz + econ) if biz is not None and econ is not None else None

        note  = ''
        if biz is not None and econ is not None:
            note = f'{int(biz)} biz / {int(econ)} econ pax'

        rows.append({
            'Airline': 'ET', 'Date': date, 'Flight / Invoice': fi,
            'Food Revenue': food, 'Service Revenue': svc, 'Total Revenue': total,
            'Pax Count': pax, 'Status': 'Invoiced', 'Notes': note,
            'Source File': path.name,
        })
    return rows


# ─── SAS Scandinavian ────────────────────────────────────────────────────────
def parse_sk(path: Path) -> list[dict]:
    print("  Parsing as SAS Scandinavian (SK) invoice CSV…")
    # Header is often on row 3 (0-indexed row 2)
    for skip in [0, 1, 2, 3]:
        try:
            df = pd.read_csv(path, skiprows=skip, dtype=str, encoding='utf-8', on_bad_lines='skip')
            if 'Inv #' in df.columns or any('inv' in c.lower() for c in df.columns):
                break
        except Exception:
            continue

    def find_col(keywords):
        for c in df.columns:
            if all(k.lower() in str(c).lower() for k in keywords):
                return c
        return None

    inv_col    = find_col(['inv', '#']) or find_col(['inv'])
    date_col   = find_col(['flt', 'date']) or find_col(['srv', 'date']) or find_col(['date'])
    total_col  = find_col(['total', 'payment']) or find_col(['payment'])
    status_col = find_col(['status'])

    if not inv_col or not total_col:
        print("  ✗  Could not find required columns in SK CSV — skipping.")
        return []

    df[total_col] = pd.to_numeric(df[total_col].str.replace(r'[$,]', '', regex=True), errors='coerce')

    # Aggregate by invoice number
    agg_cols = {total_col: 'sum'}
    if date_col:
        agg_cols[date_col] = 'first'
    if status_col:
        agg_cols[status_col] = 'first'

    grouped = df.groupby(inv_col).agg(agg_cols).reset_index()

    rows = []
    for _, r in grouped.iterrows():
        inv    = str(r[inv_col]).strip()
        date   = _to_date(r.get(date_col, '')) if date_col else ''
        total  = _money(r[total_col])
        status = str(r.get(status_col, 'PA')).strip() if status_col else 'PA'

        rows.append({
            'Airline': 'SK', 'Date': date, 'Flight / Invoice': inv,
            'Food Revenue': None, 'Service Revenue': None, 'Total Revenue': total,
            'Pax Count': None, 'Status': status,
            'Notes': 'Raw SAS CSV — food/svc split via Sage Intacct',
            'Source File': path.name,
        })
    return rows


# ─── American Airlines (Sage Intacct export) ─────────────────────────────────
def parse_aa(path: Path) -> list[dict]:
    print("  Parsing as American Airlines (AA) Sage export…")
    df = pd.read_excel(path, dtype=str)

    def find_col(keywords):
        for c in df.columns:
            if all(k.lower() in c.lower() for k in keywords):
                return c
        return None

    date_col  = find_col(['date'])
    food_col  = find_col(['food'])
    svc_col   = find_col(['service'])
    total_col = find_col(['total'])
    inv_col   = find_col(['invoice']) or find_col(['inv'])

    rows = []
    for _, r in df.iterrows():
        date  = _to_date(r.get(date_col, ''))
        if not date:
            continue
        food  = _money(r.get(food_col))
        svc   = _money(r.get(svc_col))
        total = _money(r.get(total_col))
        if total is None and food is not None and svc is not None:
            total = food + svc
        inv   = str(r.get(inv_col, '')).strip() if inv_col else ''

        rows.append({
            'Airline': 'AA', 'Date': date, 'Flight / Invoice': inv,
            'Food Revenue': food, 'Service Revenue': svc, 'Total Revenue': total,
            'Pax Count': None, 'Status': 'Invoiced', 'Notes': '',
            'Source File': path.name,
        })
    return rows


PARSERS = {'VS': parse_vs, 'ET': parse_et, 'SK': parse_sk, 'AA': parse_aa}


# ══════════════════════════════════════════════════════════════════════════════
# DEDUPLICATION
# ══════════════════════════════════════════════════════════════════════════════
def make_key(al, date, fi):
    return f"{al}|{date}|{fi}".upper()


def load_existing_keys() -> set[str]:
    if not EXCEL_FILE.exists():
        return set()
    wb  = load_workbook(EXCEL_FILE, data_only=True)
    ws  = wb[FLIGHTS_SHEET]
    keys = set()
    # Find header row
    for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
        if row[0] and str(row[0]).strip() == 'Airline':
            break
    # Read data rows (col A=airline, B=date, C=flight/invoice)
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        al = str(row[0] or '').strip()
        if not al or al == 'Airline':
            continue
        dt = row[1]
        if hasattr(dt, 'strftime'):
            dt = dt.strftime('%Y-%m-%d')
        else:
            dt = str(dt or '')[:10]
        fi = str(row[2] or '').strip()
        keys.add(make_key(al, dt, fi))
    wb.close()
    return keys


# ══════════════════════════════════════════════════════════════════════════════
# APPEND ROWS TO EXCEL
# ══════════════════════════════════════════════════════════════════════════════
def append_to_excel(new_rows: list[dict]) -> int:
    wb = load_workbook(EXCEL_FILE)
    ws = wb[FLIGHTS_SHEET]

    # Find last data row
    last_row = ws.max_row
    # Extend rows if needed
    start_row = last_row + 1

    for offset, rec in enumerate(new_rows):
        r = start_row + offset
        al   = rec['Airline']
        dt   = rec['Date']
        fi   = rec['Flight / Invoice']
        food = rec['Food Revenue']
        svc  = rec['Service Revenue']
        tot  = rec['Total Revenue']
        pax  = rec['Pax Count']
        bg   = AIRLINE_COLORS.get(al, 'FFFFFF')

        def wc(col, val, fmt=None, fg='000000', bold=False, align='left'):
            c = ws.cell(r, col, val)
            c.font      = Font(name='Arial', size=9, color=fg, bold=bold)
            c.fill      = fill(bg)
            c.border    = thin_border()
            c.alignment = Alignment(horizontal=align, vertical='center')
            if fmt:
                c.number_format = fmt
            return c

        # Parse date to Python date object if it's a string
        dt_obj = None
        if dt:
            try:
                from datetime import date as _date
                parts = dt.split('-')
                dt_obj = _date(int(parts[0]), int(parts[1]), int(parts[2]))
            except Exception:
                dt_obj = dt

        wc(1,  al,     fg='000000', bold=True, align='center')
        wc(2,  dt_obj, fmt='MMM D, YYYY', align='center')
        wc(3,  fi)
        wc(4,  food,   fmt=money_fmt(), align='right', fg='0000FF')
        wc(5,  svc,    fmt=money_fmt(), align='right', fg='0000FF')

        # Total — formula if food+svc present, else hardcoded value
        if food is not None and svc is not None:
            c6 = ws.cell(r, 6, f'=IFERROR(D{r}+E{r},"")')
        else:
            c6 = ws.cell(r, 6, tot)
        c6.font      = Font(name='Arial', size=9, color='000000')
        c6.fill      = fill(bg)
        c6.border    = thin_border()
        c6.number_format = money_fmt()
        c6.alignment = Alignment(horizontal='right', vertical='center')

        wc(7,  pax,              fmt='#,##0;-;"-"', align='right', fg='0000FF')
        wc(8,  rec['Status'],    align='center')
        wc(9,  rec['Notes'])
        wc(10, rec['Source File'], fg='888888')

        # Unique Key formula
        ck = ws.cell(r, 11, f'=A{r}&"|"&TEXT(B{r},"YYYY-MM-DD")&"|"&C{r}')
        ck.font      = Font(name='Arial', size=8, color='94A3B8', italic=True)
        ck.fill      = fill('F8FAFC')
        ck.border    = thin_border()
        ck.alignment = Alignment(horizontal='center', vertical='center')

        # Duplicate flag formula
        cd = ws.cell(r, 12,
            f'=IF(COUNTIFS($A$2:A{r},$A{r},$B$2:$B{r},$B{r},$C$2:$C{r},$C{r})>1,"⚠ DUPLICATE","✓ OK")')
        cd.font      = Font(name='Arial', size=9, bold=True)
        cd.fill      = fill(bg)
        cd.border    = thin_border()
        cd.alignment = Alignment(horizontal='center', vertical='center')

        ws.row_dimensions[r].height = 16

    wb.save(EXCEL_FILE)
    return len(new_rows)


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    print(f"\n{'═'*60}")
    print("  Sullivan Rd — Invoice Merge")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"{'═'*60}\n")

    files = [f for f in DROP_DIR.iterdir()
             if f.is_file() and f.suffix.lower() in ('.xlsx', '.csv', '.xls')
             and f.parent == DROP_DIR]

    if not files:
        print("  No new files found in Raw Data/New Invoices/")
        print("  Drop invoice files there and run again.\n")
        return

    print(f"  Found {len(files)} file(s) to process.\n")

    existing_keys = load_existing_keys()
    print(f"  Existing rows in Sullivan_Financials.xlsx: {len(existing_keys)}\n")

    total_added   = 0
    total_skipped = 0
    processed     = []

    for path in files:
        print(f"  ▶  {path.name}")
        airline = detect_airline(path)
        if not airline:
            print(f"  ✗  Could not detect airline for {path.name} — skipping.\n"
                  "     Rename the file to start with VS_, AA_, ET_, or SK_ and retry.\n")
            continue

        print(f"     Detected airline: {airline}")
        parser = PARSERS.get(airline)
        if not parser:
            print(f"  ✗  No parser for {airline} — skipping.\n")
            continue

        raw_rows = parser(path)
        if not raw_rows:
            print(f"  ✗  No data extracted from {path.name}\n")
            continue

        new_rows = []
        dups     = 0
        for rec in raw_rows:
            key = make_key(rec['Airline'], rec['Date'], rec['Flight / Invoice'])
            if key in existing_keys:
                dups += 1
            else:
                new_rows.append(rec)
                existing_keys.add(key)  # prevent intra-batch duplicates too

        print(f"     Parsed {len(raw_rows)} rows — {len(new_rows)} new, {dups} already exist")

        if new_rows:
            added = append_to_excel(new_rows)
            total_added   += added
            total_skipped += dups
            processed.append(path)
            print(f"     ✓  Added {added} rows to Sullivan_Financials.xlsx")
        else:
            print(f"     ✓  All rows already in Excel — nothing to add")
            processed.append(path)

        print()

    # Move processed files to archive
    for path in processed:
        dest = PROCESSED_DIR / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{path.name}"
        shutil.move(str(path), str(dest))
        print(f"  → Archived: {path.name}")

    print(f"\n{'─'*60}")
    print(f"  DONE — {total_added} rows added, {total_skipped} duplicates skipped")
    print(f"  Sullivan_Financials.xlsx is ready to commit and push.")
    print(f"{'─'*60}\n")


if __name__ == '__main__':
    main()
